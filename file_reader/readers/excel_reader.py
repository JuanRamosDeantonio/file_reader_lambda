import openpyxl
import xlrd
import os
import logging
import re
from pathlib import Path
from typing import Dict, List, Any, Optional
from datetime import datetime

from file_reader.readers.base_reader import BaseReader
from file_reader.core.plugin_registry import PluginRegistry
from file_reader.core.enums import OutputFormat

logger = logging.getLogger(__name__)

def escape_markdown(text: str) -> str:
    """
    Escapa caracteres especiales de markdown para evitar problemas de formato.
    """
    if not isinstance(text, str):
        text = str(text)
    
    # Escapar caracteres especiales de markdown
    text = text.replace('\\', '\\\\')  # Backslash primero
    text = text.replace('|', '\\|')    # Pipes en tablas
    text = text.replace('*', '\\*')    # Asteriscos
    text = text.replace('_', '\\_')    # Underscores
    text = text.replace('#', '\\#')    # Headers
    text = text.replace('`', '\\`')    # Code blocks
    text = text.replace('[', '\\[')    # Links
    text = text.replace(']', '\\]')    # Links
    text = text.replace('~', '\\~')    # Strikethrough
    
    return text

def escape_markdown_table_cell(text: str, max_length: int = 100) -> str:
    """
    Escapa específicamente para celdas de tabla markdown con límite configurable.
    """
    if not isinstance(text, str):
        text = str(text)
    
    # Para tablas, solo necesitamos escapar pipes y newlines
    text = text.replace('|', '\\|')
    text = text.replace('\n', ' ')  # Convertir saltos de línea a espacios
    text = text.replace('\r', ' ')
    text = text.replace('"', '\\"')  # Escapar comillas dobles
    
    # Limpiar espacios múltiples
    text = re.sub(r'\s+', ' ', text.strip())
    
    # Limitar longitud - MEJORADO: más flexible
    if len(text) > max_length:
        # Para contenido importante (que no sean solo guiones), ser más generoso
        if text.strip() != "-" and len(text) > 20:
            text = text[:max_length-3] + "..."
        else:
            text = text[:max_length]
    
    return text

def create_markdown_table(headers: list, rows: list, max_cell_length: int = 100) -> str:
    """
    Crea una tabla markdown válida con escaping apropiado y límite configurable.
    """
    if not headers or not rows:
        return ""
    
    # Escapar headers - más generoso para headers importantes
    escaped_headers = [escape_markdown_table_cell(h, max_length=150) for h in headers]
    
    # Crear tabla
    result = []
    result.append("| " + " | ".join(escaped_headers) + " |")
    result.append("| " + " | ".join("---" for _ in escaped_headers) + " |")
    
    # Escapar y agregar filas
    for row in rows:
        # Asegurar que la fila tenga el mismo número de columnas
        padded_row = row + [""] * (len(headers) - len(row))
        escaped_row = [escape_markdown_table_cell(cell, max_cell_length) if cell != "" else "-" for cell in padded_row[:len(headers)]]
        result.append("| " + " | ".join(escaped_row) + " |")
    
    return "\n".join(result)

def generate_smart_headers(first_row: list, max_cols: int = 10) -> list:
    """
    Genera headers inteligentes basados en el contenido de la primera fila.
    """
    headers = []
    for i, cell in enumerate(first_row[:max_cols]):
        if cell and str(cell).strip():
            # Usar el contenido de la celda como header si es descriptivo
            cell_str = str(cell).strip()
            if len(cell_str) > 2 and not cell_str.isdigit():
                headers.append(cell_str)
            else:
                headers.append(f"Col{i+1}")
        else:
            headers.append(f"Col{i+1}")
    
    return headers

@PluginRegistry.register("xlsx")
@PluginRegistry.register("xls")
@PluginRegistry.register("xlsm")
class ExcelReader(BaseReader):
    """
    Lector de archivos Excel optimizado para IA usando openpyxl y xlrd.
    Sin dependencias de pandas para compatibilidad con Lambda.
    Genera markdown válido con escaping automático mejorado.
    """
    
    def read(self, file_path: str) -> str:
        """
        Lee y procesa un archivo Excel con estrategias adaptativas sin pandas.
        """
        try:
            # Detectar tipo de archivo y estrategia óptima
            file_extension = Path(file_path).suffix.lower()
            strategy = self._select_reading_strategy(file_path, file_extension)
            
            logger.info(f"Procesando Excel con estrategia: {strategy}")
            
            # Extraer datos usando la estrategia seleccionada
            excel_data = self._extract_excel_data(file_path, strategy)
            
            # Formatear según configuración
            if self.config.output_format == OutputFormat.MARKDOWN_AI:
                content = self._format_ai_excel(excel_data, file_path)
            else:
                content = self._format_standard_excel(excel_data, file_path)
            
            # Aplicar formateo IA si está habilitado
            final_content = self._apply_ai_formatting(content, file_path, "excel_workbook")
            
            # Aplicar formateo JSON estructurado si es necesario
            return self._format_as_json(final_content, file_path, "excel_workbook")
            
        except Exception as e:
            logger.error(f"Error procesando archivo Excel: {e}")
            return self._handle_excel_error(file_path, str(e))
    
    def _select_reading_strategy(self, file_path: str, file_extension: str) -> str:
        """
        Selecciona la estrategia óptima de lectura basada en el archivo.
        Solo openpyxl y xlrd (sin pandas).
        """
        # Archivos .xls legacy requieren xlrd
        if file_extension == '.xls':
            return 'xlrd'
        
        # Todos los demás usan openpyxl
        else:
            return 'openpyxl'
    
    def _extract_excel_data(self, file_path: str, strategy: str) -> Dict[str, Any]:
        """
        Extrae datos usando la estrategia seleccionada.
        """
        try:
            if strategy == 'openpyxl':
                return self._read_with_openpyxl(file_path)
            elif strategy == 'xlrd':
                return self._read_with_xlrd(file_path)
            else:
                raise ValueError(f"Estrategia desconocida: {strategy}")
                
        except Exception as e:
            logger.warning(f"Estrategia {strategy} falló: {e}, intentando fallback")
            return self._read_with_fallback(file_path)
    
    def _read_with_openpyxl(self, file_path: str) -> Dict[str, Any]:
        """
        Lee Excel con openpyxl para máximo detalle y metadatos.
        """
        workbook = openpyxl.load_workbook(file_path, data_only=True)
        
        excel_data = {
            'strategy': 'openpyxl',
            'metadata': self._extract_workbook_metadata(workbook, file_path),
            'sheets': {}
        }
        
        # Procesar cada hoja
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            
            # Extraer datos de la hoja
            sheet_data = []
            max_rows_to_process = min(sheet.max_row, 1000)  # Límite para archivos grandes
            
            for row in sheet.iter_rows(max_row=max_rows_to_process, values_only=True):
                # Filtrar filas completamente vacías
                if any(cell is not None for cell in row):
                    sheet_data.append([cell if cell is not None else "" for cell in row])
            
            if sheet_data:
                excel_data['sheets'][sheet_name] = {
                    'data': sheet_data,
                    'dimensions': f"{sheet.max_row} x {sheet.max_column}",
                    'processed_rows': len(sheet_data),
                    'has_merged_cells': len(sheet.merged_cells.ranges) > 0,
                    'formulas': self._extract_formulas(sheet),
                    'data_types': self._analyze_sheet_types(sheet_data)
                }
                
                if sheet.max_row > 1000:
                    excel_data['sheets'][sheet_name]['note'] = f"Mostrando primeras 1000 filas de {sheet.max_row} total"
        
        workbook.close()
        return excel_data
    
    def _read_with_xlrd(self, file_path: str) -> Dict[str, Any]:
        """
        Lee archivos .xls legacy con xlrd.
        """
        workbook = xlrd.open_workbook(file_path)
        
        excel_data = {
            'strategy': 'xlrd',
            'metadata': {
                'file_name': Path(file_path).name,
                'file_type': 'Legacy Excel (.xls)',
                'file_size_mb': round(os.path.getsize(file_path) / (1024 * 1024), 2),
                'sheet_names': workbook.sheet_names(),
                'total_sheets': workbook.nsheets
            },
            'sheets': {}
        }
        
        # Procesar cada hoja
        for sheet_name in workbook.sheet_names():
            sheet = workbook.sheet_by_name(sheet_name)
            
            sheet_data = []
            max_rows_to_process = min(sheet.nrows, 1000)  # Límite para archivos grandes
            
            for row_idx in range(max_rows_to_process):
                row_data = []
                for col_idx in range(sheet.ncols):
                    cell = sheet.cell(row_idx, col_idx)
                    row_data.append(self._convert_xlrd_cell_value(cell))
                
                # Solo agregar filas que no estén completamente vacías
                if any(str(cell).strip() for cell in row_data):
                    sheet_data.append(row_data)
            
            excel_data['sheets'][sheet_name] = {
                'data': sheet_data,
                'dimensions': f"{sheet.nrows} x {sheet.ncols}",
                'processed_rows': len(sheet_data),
                'data_types': self._analyze_sheet_types(sheet_data)
            }
            
            if sheet.nrows > 1000:
                excel_data['sheets'][sheet_name]['note'] = f"Mostrando primeras 1000 filas de {sheet.nrows} total"
        
        return excel_data
    
    def _read_with_fallback(self, file_path: str) -> Dict[str, Any]:
        """
        Estrategia de fallback cuando otras fallan.
        """
        file_extension = Path(file_path).suffix.lower()
        fallback_strategies = []
        
        if file_extension == '.xls':
            fallback_strategies = ['xlrd', 'openpyxl']
        else:
            fallback_strategies = ['openpyxl', 'xlrd']
        
        for strategy in fallback_strategies:
            try:
                logger.info(f"Intentando estrategia de fallback: {strategy}")
                
                if strategy == 'openpyxl':
                    return self._read_with_openpyxl(file_path)
                elif strategy == 'xlrd':
                    return self._read_with_xlrd(file_path)
                    
            except Exception as e:
                logger.warning(f"Fallback {strategy} falló: {e}")
                continue
        
        # Si todo falla, retornar estructura mínima
        return {
            'strategy': 'fallback_failed',
            'metadata': {
                'file_name': Path(file_path).name,
                'error': 'Todas las estrategias de lectura fallaron'
            },
            'sheets': {}
        }
    
    def _extract_workbook_metadata(self, workbook: openpyxl.Workbook, file_path: str) -> Dict[str, Any]:
        """
        Extrae metadatos completos del workbook.
        """
        props = workbook.properties
        
        return {
            'file_name': Path(file_path).name,
            'file_size_mb': round(os.path.getsize(file_path) / (1024 * 1024), 2),
            'title': props.title or 'Sin título',
            'creator': props.creator or 'Desconocido',
            'created': props.created.isoformat() if props.created else None,
            'modified': props.modified.isoformat() if props.modified else None,
            'subject': props.subject or '',
            'description': props.description or '',
            'sheet_names': workbook.sheetnames,
            'total_sheets': len(workbook.sheetnames),
            'defined_names': [name.name for name in workbook.defined_names] if workbook.defined_names else [],
            'has_vba': workbook.vba_archive is not None
        }
    
    def _extract_formulas(self, sheet) -> List[Dict[str, str]]:
        """
        Extrae fórmulas de una hoja.
        """
        formulas = []
        
        # Solo revisar primeras 100 filas para optimización
        max_rows = min(sheet.max_row, 100)
        max_cols = min(sheet.max_column, 50)
        
        for row_num in range(1, max_rows + 1):
            for col_num in range(1, max_cols + 1):
                cell = sheet.cell(row=row_num, column=col_num)
                if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
                    formulas.append({
                        'cell': cell.coordinate,
                        'formula': cell.value
                    })
                    
                    if len(formulas) >= 10:  # Límite de 10 fórmulas
                        return formulas
        
        return formulas
    
    def _analyze_sheet_types(self, sheet_data: List[List]) -> Dict[str, int]:
        """
        Analiza tipos de datos en los datos de la hoja.
        """
        type_counts = {
            'text': 0,
            'number': 0,
            'date': 0,
            'boolean': 0,
            'empty': 0
        }
        
        # Analizar primeras 100 celdas para detectar tipos
        cells_analyzed = 0
        for row in sheet_data[:20]:  # Primeras 20 filas
            for cell in row[:10]:  # Primeras 10 columnas
                cells_analyzed += 1
                
                if cell == "" or cell is None:
                    type_counts['empty'] += 1
                elif isinstance(cell, bool):
                    type_counts['boolean'] += 1
                elif isinstance(cell, (int, float)):
                    type_counts['number'] += 1
                elif isinstance(cell, datetime):
                    type_counts['date'] += 1
                else:
                    # Intentar detectar fechas en strings
                    cell_str = str(cell).strip()
                    if self._looks_like_date(cell_str):
                        type_counts['date'] += 1
                    else:
                        type_counts['text'] += 1
                
                if cells_analyzed >= 100:
                    break
            if cells_analyzed >= 100:
                break
        
        return type_counts
    
    def _looks_like_date(self, text: str) -> bool:
        """
        Detecta si un texto parece una fecha.
        """
        if len(text) < 6:
            return False
        
        date_indicators = ['-', '/', ':', 'T', '202', '201', '199']
        return any(indicator in text for indicator in date_indicators)
    
    def _convert_xlrd_cell_value(self, cell):
        """
        Convierte valores de celda xlrd a formato apropiado.
        """
        if cell.ctype == xlrd.XL_CELL_DATE:
            try:
                return xlrd.xldate_as_datetime(cell.value, 0).strftime('%Y-%m-%d %H:%M:%S')
            except:
                return str(cell.value)
        elif cell.ctype == xlrd.XL_CELL_NUMBER:
            # Convertir a int si es número entero
            if cell.value == int(cell.value):
                return int(cell.value)
            return cell.value
        elif cell.ctype == xlrd.XL_CELL_BOOLEAN:
            return bool(cell.value)
        else:
            return str(cell.value) if cell.value else ""
    
    def _format_ai_excel(self, excel_data: Dict[str, Any], file_path: str) -> str:
        """
        Formatea datos de Excel optimizado para IA con markdown válido mejorado.
        """
        metadata = excel_data['metadata']
        
        # Header con análisis del archivo (escapando nombres y datos)
        filename = escape_markdown(metadata['file_name'])
        strategy = escape_markdown(excel_data['strategy'])
        creator = escape_markdown(metadata.get('creator', 'Desconocido'))
        
        # Construir header usando concatenación de strings en lugar de f-string multilínea
        header_lines = [
            "## 📊 Excel Workbook Analysis",
            f"- **Archivo:** {filename}",
            f"- **Estrategia de lectura:** {strategy} (sin pandas)",
            f"- **Tamaño:** {metadata.get('file_size_mb', 'Desconocido')} MB",
            f"- **Hojas:** {metadata.get('total_sheets', len(excel_data['sheets']))}",
            f"- **Creado por:** {creator}",
            f"- **Fecha creación:** {metadata.get('created', 'Desconocida')}",
            "",
            ""
        ]
        
        header = "\n".join(header_lines)
        
        # Mostrar información de hojas con escaping apropiado
        if excel_data['sheets']:
            header += "### 📋 Estructura de Hojas\n"
            for sheet_name, sheet_info in excel_data['sheets'].items():
                escaped_sheet_name = escape_markdown(sheet_name)
                if 'error' in sheet_info:
                    escaped_error = escape_markdown(sheet_info['error'])
                    header += f"- **{escaped_sheet_name}**: ❌ Error - {escaped_error}\n"
                else:
                    dimensions = escape_markdown(sheet_info.get('dimensions', 'Desconocido'))
                    processed = sheet_info.get('processed_rows', 0)
                    header += f"- **{escaped_sheet_name}**: {dimensions} ({processed} filas procesadas)"
                    
                    if 'data_types' in sheet_info:
                        types = sheet_info['data_types']
                        main_types = [k for k, v in types.items() if v > 0 and k != 'empty']
                        if main_types:
                            types_text = escape_markdown(', '.join(main_types))
                            header += f" (Tipos: {types_text})"
                    
                    if 'note' in sheet_info:
                        note_text = escape_markdown(sheet_info['note'])
                        header += f" - {note_text}"
                    
                    header += "\n"
            header += "\n"
        
        # Contenido de las hojas con tablas markdown válidas MEJORADAS
        content_sections = []
        
        for sheet_name, sheet_info in excel_data['sheets'].items():
            if 'error' in sheet_info:
                continue
                
            sheet_data = sheet_info['data']
            if not sheet_data:
                continue
            
            escaped_sheet_name = escape_markdown(sheet_name)
            section = f"### 📄 Hoja: {escaped_sheet_name}\n"
            
            # Información adicional de la hoja con escaping
            if 'formulas' in sheet_info and sheet_info['formulas']:
                section += f"**Fórmulas encontradas:** {len(sheet_info['formulas'])}\n"
                for formula in sheet_info['formulas'][:3]:  # Mostrar primeras 3
                    cell_ref = escape_markdown(formula['cell'])
                    formula_text = escape_markdown(formula['formula'])
                    section += f"- {cell_ref}: `{formula_text}`\n"
                section += "\n"
            
            # Crear tabla markdown válida MEJORADA
            if len(sheet_data) > 0:
                # Limitar a primeras 15 filas para más contenido
                preview_data = sheet_data[:16] if len(sheet_data) > 15 else sheet_data
                
                if len(preview_data) > 0:
                    # MEJORADO: Mejor detección de headers
                    if self._looks_like_headers(preview_data[0]):
                        headers = generate_smart_headers(preview_data[0], max_cols=12)
                        rows = preview_data[1:15]  # Siguientes 14 filas
                    else:
                        # Headers más informativos basados en contenido
                        max_cols = max(len(row) for row in preview_data) if preview_data else 0
                        headers = generate_smart_headers(preview_data[0] if preview_data else [], min(max_cols, 12))
                        rows = preview_data[:15]
                    
                    # MEJORADO: Más columnas permitidas (12 en lugar de 10)
                    if len(headers) > 12:
                        headers = headers[:12]
                        rows = [row[:12] for row in rows]
                    
                    # Crear tabla con escaping apropiado y límites más generosos
                    if headers and rows:
                        table = create_markdown_table(headers, rows, max_cell_length=120)
                        section += table + "\n"
                        
                        if len(sheet_data) > 15:
                            section += f"\n*Mostrando primeras 15 filas de {len(sheet_data)} total*\n"
                    else:
                        section += "*No hay datos para mostrar*\n"
            
            content_sections.append(section)
        
        return header + "\n".join(content_sections)
    
    def _looks_like_headers(self, row: List) -> bool:
        """
        Determina si una fila parece contener headers - MEJORADO.
        """
        if not row:
            return False
        
        # Headers típicamente son texto, no números
        text_count = 0
        meaningful_content = 0
        
        for cell in row:
            if cell and str(cell).strip():
                cell_str = str(cell).strip()
                meaningful_content += 1
                
                # Es probable que sea header si:
                # - Es texto (no solo números)
                # - Tiene longitud razonable
                # - No es solo un símbolo
                if (isinstance(cell, str) and len(cell_str) > 1 and 
                    not cell_str.isdigit() and cell_str != "-"):
                    text_count += 1
        
        # Es header si 60% o más son texto significativo
        return meaningful_content > 0 and text_count >= meaningful_content * 0.6
    
    def _format_standard_excel(self, excel_data: Dict[str, Any], file_path: str) -> str:
        """
        Formatea Excel en formato estándar con markdown válido.
        """
        if self.config.output_format == OutputFormat.PLAIN:
            # Formato texto plano (no necesita escaping)
            result = []
            for sheet_name, sheet_info in excel_data['sheets'].items():
                if 'data' in sheet_info:
                    result.append(f"=== {sheet_name} ===")
                    for row in sheet_info['data'][:50]:  # Primeras 50 filas
                        result.append("\t".join(str(cell) for cell in row))
                    result.append("")
            return "\n".join(result)
        else:
            # Formato markdown estándar con escaping MEJORADO
            result = []
            for sheet_name, sheet_info in excel_data['sheets'].items():
                if 'data' in sheet_info and sheet_info['data']:
                    escaped_sheet_name = escape_markdown(sheet_name)
                    result.append(f"## {escaped_sheet_name}")
                    
                    data = sheet_info['data']
                    if len(data) > 0:
                        # Usar primera fila como headers con generación inteligente
                        headers = generate_smart_headers(data[0], max_cols=12) if data else []
                        rows = data[1:8] if len(data) > 1 else []  # Más filas (8 en lugar de 5)
                        
                        if headers and rows:
                            # Limitar columnas para evitar tablas muy anchas
                            max_cols = min(len(headers), 12)
                            headers = headers[:max_cols]
                            rows = [row[:max_cols] for row in rows]
                            
                            # Crear tabla markdown válida con límites más generosos
                            table = create_markdown_table(headers, rows, max_cell_length=120)
                            result.append(table)
                        elif headers:
                            # Solo headers, sin datos
                            table = create_markdown_table(headers, [], max_cell_length=120)
                            result.append(table)
                    
                    result.append("")
            
            return "\n".join(result)
    
    def _handle_excel_error(self, file_path: str, error_message: str) -> str:
        """
        Maneja errores de lectura de Excel con información útil y markdown válido.
        """
        file_name = escape_markdown(Path(file_path).name)
        escaped_error = escape_markdown(error_message)
        file_extension = escape_markdown(Path(file_path).suffix)
        
        # Construir mensaje de error usando concatenación
        error_lines = [
            "## ❌ Error procesando archivo Excel",
            "",
            f"**Archivo:** {file_name}",
            f"**Error:** {escaped_error}",
            "",
            "### 💡 Posibles soluciones:",
            "- Verificar que el archivo no esté corrupto",
            "- Asegurar que el archivo no esté protegido con contraseña",
            "- Verificar que sea un archivo Excel válido (.xlsx, .xls, .xlsm)",
            "- Intentar abrir el archivo en Excel para verificar su integridad",
            "",
            "### 📋 Información técnica:",
            f"- **Extensión detectada:** {file_extension}",
            f"- **Tamaño de archivo:** {round(os.path.getsize(file_path) / (1024 * 1024), 2)} MB",
            "- **Método de lectura:** openpyxl + xlrd (sin pandas)"
        ]
        
        return "\n".join(error_lines)

def validate_markdown_output(markdown_text: str) -> dict:
    """
    Valida si el markdown generado es sintácticamente correcto.
    """
    issues = []
    
    lines = markdown_text.split('\n')
    in_table = False
    table_columns = 0
    
    for i, line in enumerate(lines, 1):
        # Verificar tablas
        if '|' in line and not line.strip().startswith('```'):
            # Contar columnas en la tabla
            pipes = line.count('|') - line.count('\\|')
            
            if not in_table:
                in_table = True
                table_columns = pipes
            else:
                # Verificar consistencia de columnas
                if pipes != table_columns and pipes > 0:
                    issues.append(f"Línea {i}: Inconsistencia en número de columnas de tabla")
            
            # Verificar que los pipes estén balanceados
            if pipes < 2:
                issues.append(f"Línea {i}: Tabla con formato incorrecto (pipes insuficientes)")
        
        # Verificar separadores de tabla
        elif in_table and '---' in line:
            continue  # Separador válido
        elif in_table and '|' not in line and line.strip():
            in_table = False
            table_columns = 0
        
        # Verificar headers
        if line.startswith('#'):
            header_level = len(line) - len(line.lstrip('#'))
            if header_level > 6:
                issues.append(f"Línea {i}: Header nivel {header_level} no válido (máximo 6)")
            
            # Verificar que hay espacio después del #
            if len(line) > header_level and line[header_level] != ' ':
                issues.append(f"Línea {i}: Falta espacio después de # en header")
        
        # Verificar bloques de código
        if line.strip().startswith('```'):
            # Verificar que el bloque de código esté bien formado
            if line.count('```') == 1:
                continue  # Inicio o fin de bloque
            else:
                issues.append(f"Línea {i}: Bloque de código mal formado")
    
    return {
        'valid': len(issues) == 0,
        'issues': issues,
        'total_lines': len(lines),
        'tables_found': markdown_text.count('|') > 0
    }